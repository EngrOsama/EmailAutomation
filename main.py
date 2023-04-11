import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook

# load list of emails from Excel file
wb = load_workbook('email_list.xlsx')
ws = wb.active
emails = [cell for cell in ws['A'] if cell.offset(0, 1).value != 'done'] # check if neighboring cell has "done" written in it
emails = emails[:400] # limit the number of emails to be sent to <=500

# set up SMTP server and login
smtp_server = 'smtp.gmail.com'
smtp_port = 587
smtp_username = 'your_email@gmail.com'
smtp_password = 'your_email_password'
sender_email = smtp_username

# loop through list of emails and send message with attachment
for counter, email_cell in enumerate(emails):
    email = email_cell.value

    # create message container
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = email
    msg['Subject'] = 'Your Subject'

    # attach PDF file
    filename = 'your_file.pdf'
    attachment = open(filename, 'rb')
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(part)

    # add message body
    body = 'Your message body here'
    msg.attach(MIMEText(body, 'plain'))

    # connect to SMTP server and send message
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_username, smtp_password)
        text = msg.as_string()
        server.sendmail(sender_email, email, text)
        server.quit()

        # add delay timer after sending each email
        time.sleep(0) # set delay in seconds

        # write 'done' in the next column
        done_cell = email_cell.offset(0, 1)
        done_cell.value = 'done'

        # save the file
        wb.save('email_list.xlsx')
        print(f"{counter+1}. Email sent to {email}")

    except smtplib.SMTPRecipientsRefused:
        print(f"Invalid email address: {email}")
        continue
